import logging
import os
from pathlib import Path

from dotenv import load_dotenv

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)s %(name)s %(message)s",
)

_BASE_DIR = Path(__file__).resolve().parent.parent

load_dotenv()

# Application Insights — must be configured before other imports to instrument all libraries
_ai_connection_string = os.getenv("APPLICATIONINSIGHTS_CONNECTION_STRING")
if _ai_connection_string:
    try:
        from azure.monitor.opentelemetry import configure_azure_monitor
        configure_azure_monitor(connection_string=_ai_connection_string)
    except Exception as _ai_exc:
        logging.getLogger(__name__).warning("configure_azure_monitor failed, telemetry disabled: %s", _ai_exc)

from fastapi import FastAPI
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, RedirectResponse
from fastapi.staticfiles import StaticFiles

from app.routers import basket, bulk_pricing, chat, diagram, vm_prices, sql_prices

app = FastAPI(title="Azure VM Pricing Bot", version="1.0.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

app.mount("/static", StaticFiles(directory=str(_BASE_DIR / "static")), name="static")

app.include_router(chat.router, prefix="/api")
app.include_router(basket.router, prefix="/api/basket")
app.include_router(diagram.router, prefix="/api/diagram")
app.include_router(vm_prices.router, prefix="/api/vm-prices")
app.include_router(sql_prices.router, prefix="/api/sql-vm-prices", tags=["sql-pricing"])
app.include_router(bulk_pricing.router, prefix="/api/bulk-pricing", tags=["bulk-pricing"])


@app.get("/")
def root_redirect():
    return RedirectResponse(url="https://hyperxen.ai", status_code=301)


@app.get("/pricing")
async def pricing() -> FileResponse:
    return FileResponse(
        str(_BASE_DIR / "static" / "index.html"),
        headers={"Cache-Control": "no-cache, no-store, must-revalidate"},
    )


@app.get("/architect")
async def architect() -> FileResponse:
    return FileResponse(
        str(_BASE_DIR / "static" / "architect.html"),
        headers={"Cache-Control": "no-cache, no-store, must-revalidate"},
    )



@app.get("/compare")
async def compare() -> FileResponse:
    return FileResponse(
        str(_BASE_DIR / "static" / "compare.html"),
        headers={"Cache-Control": "no-cache, no-store, must-revalidate"},
    )


@app.get("/bulk-pricing")
async def bulk_pricing_page() -> FileResponse:
    return FileResponse(
        str(_BASE_DIR / "static" / "bulk_pricing.html"),
        headers={"Cache-Control": "no-cache, no-store, must-revalidate"},
    )


@app.get("/api/bulk-pricing/template")
async def bulk_pricing_template():
    """Serve the Excel inventory template with example rows."""
    from fastapi.responses import StreamingResponse
    from io import BytesIO
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "VM Inventory"

    HDR_FILL = PatternFill(fill_type="solid", fgColor="FF0078D4")
    HDR_FONT = Font(bold=True, color="FFFFFFFF")
    CTR = Alignment(horizontal="center")

    headers = ["Location", "VM Name", "OS", "MS SQL", "Server Role",
               "vCPUs", "Mem (GB)", "Provisioned Space (GB)",
               "Windows AHB", "SQL AHB", "Disk Type"]
    ws.append(headers)
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i)
        c.font = HDR_FONT
        c.fill = HDR_FILL
        c.alignment = CTR

    # Columns 1-8 are the original required columns.
    # Columns 9-11 are the new optional columns — blank = use the default.
    examples = [
        # All defaults — Standard SSD, License-Included, no AHB
        ["AU",  "PROD-WEB-01",  "Windows Server", "--",         "Web",      2,  8,   128,  "",    "",    ""],
        # Windows AHB: compute priced at Linux rate, no Windows OS license charge
        ["AU",  "PROD-WEB-02",  "Windows Server", "--",         "Web",      4,  16,  128,  "Yes", "",    ""],
        # SQL AHB: SQL Server license waived (BYOL)
        ["AU",  "PROD-DB-01",   "Windows Server", "Enterprise", "Database", 8,  32,  1024, "",    "Yes", ""],
        # Premium SSD (P-series) — tiered, monthly flat rate per tier
        ["AU",  "PROD-DB-02",   "Windows Server", "Standard",   "Database", 4,  16,  512,  "",    "",    "Premium SSD"],
        # Premium SSD v2 — provisioned $/GiB/month, more flexible sizing
        ["AU",  "PERF-DB-01",   "Windows Server", "--",         "Database", 8,  32,  2048, "",    "",    "Premium SSD v2"],
        # Linux default — no AHB needed, Standard SSD
        ["USA", "DR-WEB-01",    "Linux",           "--",         "Web",      2,  8,   128,  "",    "",    ""],
    ]
    for row in examples:
        ws.append(row)

    col_widths = [10, 20, 18, 14, 14, 8, 10, 22, 13, 10, 16]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    # Notes sheet
    ws2 = wb.create_sheet("Notes")
    notes = [
        ("Location codes:",          "AU = Australia East,  USA = East US"),
        ("OS values:",               "Windows Server  /  Linux"),
        ("MS SQL values:",           "--  (none)  |  Developer  |  Express  |  Web  |  Standard  |  Enterprise"),
        ("vCPUs:",                   "Integer — the VM's requested vCPU count"),
        ("Mem (GB):",                "Integer — the VM's requested RAM in GB"),
        ("Provisioned Space (GB):",  "Float — OS disk size in GiB"),
        ("Windows AHB:",             "Yes = Azure Hybrid Benefit applied — compute priced at Linux rate, no Windows OS license charge.  Blank = License-Included (default)."),
        ("SQL AHB:",                 "Yes = SQL Server license waived (BYOL).  Blank = License-Included PAYG rate (default).  Only relevant when MS SQL is set."),
        ("Disk Type:",               "Standard SSD (default)  |  Premium SSD  |  Premium SSD v2.  Blank = Standard SSD."),
        ("",                         "Standard SSD: rounded UP to nearest E-series tier, flat monthly rate."),
        ("",                         "Premium SSD: rounded UP to nearest P-series tier, flat monthly rate."),
        ("",                         "Premium SSD v2: provisioned $/GiB/month × disk size (capacity cost only)."),
    ]
    for r, (k, v) in enumerate(notes, 1):
        ws2.cell(row=r, column=1, value=k).font = Font(bold=True)
        ws2.cell(row=r, column=2, value=v)
    ws2.column_dimensions["A"].width = 26
    ws2.column_dimensions["B"].width = 90

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="vm-inventory-template.xlsx"'},
    )


@app.get("/health")
async def health() -> dict:
    return {"status": "ok", "version": "1.0.0"}
