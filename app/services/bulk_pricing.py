"""
Bulk VM pricing — Public-Cloud-VM-Listing Excel sheet processor.

Input columns:  Location | VM Name | OS | MS SQL | Server Role | vCPUs | Mem (GB) | Provisioned Space (GB)
Optional:       Windows AHB | SQL AHB | Disk Type

Reuses the same SKU-matching and pricing functions as the chat advisor path.
All pricing math flows through existing sql_pricing / azure_pricing utilities —
no duplicate formulas here.
"""

import asyncio
import logging
import math
import re
from typing import Optional

logger = logging.getLogger(__name__)

HOURS_PER_MONTH = 730

# ── Location → Azure region mapping ──────────────────────────────────────────
# Add new country codes here as needed; values are (arm_region, display_label)
REGION_MAP: dict[str, tuple[str, str]] = {
    "AU":  ("australiaeast", "Australia East"),
    "USA": ("eastus",        "East US"),
}

# ── Azure disk tier tables ────────────────────────────────────────────────────
# Each entry: (tier_name, size_gib). Size thresholds are identical across E and
# P series — only the prefix differs.
_SSD_TIERS: list[tuple[str, int]] = [
    ("E1",  4),    ("E2",  8),    ("E3",  16),   ("E4",  32),   ("E6",  64),
    ("E10", 128),  ("E15", 256),  ("E20", 512),  ("E30", 1024),
    ("E40", 2048), ("E50", 4096), ("E60", 8192), ("E70", 16384), ("E80", 32767),
]
_PREMIUM_SSD_TIERS: list[tuple[str, int]] = [
    ("P1",  4),    ("P2",  8),    ("P3",  16),   ("P4",  32),   ("P6",  64),
    ("P10", 128),  ("P15", 256),  ("P20", 512),  ("P30", 1024),
    ("P40", 2048), ("P50", 4096), ("P60", 8192), ("P70", 16384), ("P80", 32767),
]

# ── Module-level caches (lives for the duration of one process run) ───────────
# Keyed by (region, os_type, vcpus, ram_gb) → list of matched SKU dicts
_sku_match_cache: dict[tuple, list[dict]] = {}
# Keyed by (region, sku_name) → list of price item dicts from Azure Retail API
_vm_price_cache: dict[tuple, list[dict]] = {}
# Keyed by region → {tier_name: monthly_usd} for Standard SSD
_ssd_price_cache: dict[str, dict[str, float]] = {}
# Keyed by region → {tier_name: monthly_usd} for Premium SSD (P-series)
_premium_ssd_cache: dict[str, dict[str, float]] = {}
# Keyed by region → monthly USD/GiB for Premium SSD v2 capacity
_pv2_rate_cache: dict[str, Optional[float]] = {}


# ── Utility: disk tier resolution ────────────────────────────────────────────

def resolve_ssd_tier(size_gb: float) -> tuple[str, int]:
    """Round UP to the smallest Azure Standard SSD tier that fits."""
    size_ceil = math.ceil(size_gb)
    for tier_name, tier_gib in _SSD_TIERS:
        if tier_gib >= size_ceil:
            return tier_name, tier_gib
    return "E80", 32767


def resolve_premium_ssd_tier(size_gb: float) -> tuple[str, int]:
    """Round UP to the smallest Azure Premium SSD tier that fits."""
    size_ceil = math.ceil(size_gb)
    for tier_name, tier_gib in _PREMIUM_SSD_TIERS:
        if tier_gib >= size_ceil:
            return tier_name, tier_gib
    return "P80", 32767


def _resolve_disk_type(raw: str) -> str:
    """Map the 'Disk Type' column to an internal key.

    'Premium SSD v2' → 'premium_ssd_v2'
    'Premium SSD'    → 'premium_ssd'
    ''  / 'Standard SSD' / anything else → 'standard_ssd'  (default)
    """
    lower = (raw or "").strip().lower()
    if "v2" in lower:
        return "premium_ssd_v2"
    if "premium" in lower:
        return "premium_ssd"
    return "standard_ssd"


# ── Utility: SQL edition mapping ──────────────────────────────────────────────

def _resolve_sql_edition(raw: str) -> tuple[Optional[str], Optional[str], bool]:
    """Map the Excel MS SQL column value to (display_edition, billing_edition, is_warning).

    display_edition  — what to show in the output table (None = no SQL)
    billing_edition  — what to pass to sql_license_hourly() (None = no SQL)
    is_warning       — True if the value was unrecognized; caller should log a warning

    Mapping:
        "--" or ""   → (None, None, False)              — no SQL Server
        "Developer"  → ("Developer", "Express", False)  — free; display as Developer
        "Web"        → ("Web", "Web", False)
        "Standard"   → ("Standard", "Standard", False)
        "Enterprise" → ("Enterprise", "Enterprise", False)
        "Express"    → ("Express", "Express", False)
        anything else → (None, None, True)              — unrecognized; no SQL, emit warning
    """
    raw = (raw or "").strip()
    lower = raw.lower()
    if not lower or lower == "--":
        return None, None, False
    if lower == "developer":
        return "Developer", "Express", False
    _known = {"web": "Web", "standard": "Standard", "enterprise": "Enterprise", "express": "Express"}
    if lower in _known:
        edition = _known[lower]
        return edition, edition, False
    return None, None, True


def _resolve_os(raw: str) -> Optional[str]:
    """Map Excel OS column to 'Windows' or 'Linux'. Returns None if unrecognized."""
    lower = (raw or "").strip().lower()
    if "windows" in lower:
        return "Windows"
    if "linux" in lower:
        return "Linux"
    return None


# ── Excel parser ──────────────────────────────────────────────────────────────

def parse_inventory_xlsx(file_bytes: bytes) -> tuple[list[dict], list[str]]:
    """Parse a Public-Cloud-VM-Listing Excel file into a list of validated row dicts.

    Returns (rows, warnings).  Skips the Total/summary row and any rows with
    unrecognized Location codes.  Unknown OS defaults to Windows with a warning.
    Unknown MS SQL values default to no SQL with a warning.
    """
    import openpyxl
    from io import BytesIO

    warnings: list[str] = []
    rows: list[dict] = []

    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    ws = wb.active

    # ── Find header row ───────────────────────────────────────────────────────
    header_row_idx: Optional[int] = None
    headers: list[str] = []
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        cells = [str(c or "").strip() for c in row]
        if "Location" in cells and "VM Name" in cells:
            header_row_idx = row_idx
            headers = cells
            break

    if not headers:
        warnings.append("ERROR: Could not find header row with 'Location' and 'VM Name' columns")
        return [], warnings

    def _col(name: str) -> Optional[int]:
        for i, h in enumerate(headers):
            if h.strip().lower() == name.strip().lower():
                return i
        return None

    col_location = _col("Location")
    col_vmname   = _col("VM Name")
    col_os       = _col("OS")
    col_sql      = _col("MS SQL")
    col_role     = _col("Server Role")
    col_vcpus    = _col("vCPUs")
    col_mem      = _col("Mem (GB)")
    col_disk     = _col("Provisioned Space (GB)")
    # Optional new columns — None if not present in the uploaded file (backward-compat)
    col_win_ahb  = _col("Windows AHB")
    col_sql_ahb  = _col("SQL AHB")
    col_disk_type = _col("Disk Type")

    missing = [nm for nm, idx in [
        ("Location", col_location), ("VM Name", col_vmname),
        ("vCPUs",    col_vcpus),    ("Mem (GB)", col_mem),
    ] if idx is None]
    if missing:
        warnings.append(f"ERROR: Missing required columns: {', '.join(missing)}")
        return [], warnings

    # ── Iterate data rows ─────────────────────────────────────────────────────
    for sheet_row_num, row in enumerate(
        ws.iter_rows(min_row=(header_row_idx + 1), values_only=True),
        start=header_row_idx + 1,
    ):
        def _cell(col_idx: Optional[int]) -> str:
            if col_idx is None or col_idx >= len(row):
                return ""
            return str(row[col_idx] or "").strip()

        # Skip entirely blank rows
        if all(c is None or str(c).strip() == "" for c in row):
            continue

        location = _cell(col_location)
        vm_name  = _cell(col_vmname)

        # Skip Total / summary rows
        if location.lower() in ("total", "") or not vm_name:
            continue

        # Skip unrecognized location codes
        location_upper = location.upper()
        if location_upper not in REGION_MAP:
            warnings.append(
                f"Row {sheet_row_num} '{vm_name}': unknown Location '{location}' — skipped"
            )
            continue

        # Parse vCPUs
        try:
            vcpus = int(float(_cell(col_vcpus)))
            if vcpus <= 0:
                raise ValueError("non-positive")
        except (ValueError, TypeError):
            warnings.append(
                f"Row {sheet_row_num} '{vm_name}': invalid vCPUs '{_cell(col_vcpus)}' — skipped"
            )
            continue

        # Parse RAM
        try:
            ram_gb = int(float(_cell(col_mem)))
            if ram_gb <= 0:
                raise ValueError("non-positive")
        except (ValueError, TypeError):
            warnings.append(
                f"Row {sheet_row_num} '{vm_name}': invalid Mem '{_cell(col_mem)}' — skipped"
            )
            continue

        # Parse provisioned disk size
        try:
            disk_gb = float(_cell(col_disk)) if _cell(col_disk) else 0.0
        except (ValueError, TypeError):
            disk_gb = 0.0

        os_raw  = _cell(col_os)  or "Windows Server"
        sql_raw = _cell(col_sql) or "--"
        role    = _cell(col_role)

        # Resolve OS
        os_type = _resolve_os(os_raw)
        if os_type is None:
            warnings.append(
                f"Row {sheet_row_num} '{vm_name}': unrecognized OS '{os_raw}' — defaulting to Windows"
            )
            os_type = "Windows"

        # Resolve SQL edition
        sql_display, sql_billing, sql_warn = _resolve_sql_edition(sql_raw)
        if sql_warn:
            warnings.append(
                f"Row {sheet_row_num} '{vm_name}': unrecognized MS SQL value '{sql_raw}'"
                " — no SQL pricing applied"
            )

        arm_region, region_label = REGION_MAP[location_upper]
        disk_gb_eff = disk_gb if disk_gb > 0 else 128.0
        ssd_tier, ssd_gib = resolve_ssd_tier(disk_gb_eff)

        # Parse optional new columns (default to False / standard_ssd when absent)
        win_ahb   = _cell(col_win_ahb).strip().lower() == "yes"
        sql_ahb   = _cell(col_sql_ahb).strip().lower() == "yes"
        disk_type = _resolve_disk_type(_cell(col_disk_type))

        rows.append({
            "sheet_row":    sheet_row_num,
            "vm_name":      vm_name,
            "location":     location,
            "arm_region":   arm_region,
            "region_label": region_label,
            "os_type":      os_type,
            "vcpus_req":    vcpus,
            "ram_gb_req":   ram_gb,
            "disk_gb_raw":  disk_gb,
            "ssd_tier":     ssd_tier,
            "ssd_gib":      ssd_gib,
            "sql_display":  sql_display,
            "sql_billing":  sql_billing,
            "role":         role,
            "win_ahb":      win_ahb,
            "sql_ahb":      sql_ahb,
            "disk_type":    disk_type,
        })

    logger.info(
        "parse_inventory_xlsx: %d valid rows, %d warnings", len(rows), len(warnings)
    )
    return rows, warnings


# ── Cached SKU matching ───────────────────────────────────────────────────────

async def _pick_sku_cached(
    region: str, os_type: str, vcpus: int, ram_gb: int
) -> list[dict]:
    """Best-fit D/E/F/B SKU for bulk pricing via AI Search (cached by exact spec).

    Uses search_skus() which queries the vm-skus AI Search index restricted to
    D/E/F/B series only — preventing GPU/HPC/M-series SKUs from contaminating
    bulk pricing results regardless of region or ARM credential availability.
    """
    from app.agents.sku_advisor_agent import search_skus

    key = (region, os_type, vcpus, ram_gb)
    if key in _sku_match_cache:
        return _sku_match_cache[key]

    docs = await asyncio.to_thread(search_skus, {"vcpus": vcpus, "ram_gb": ram_gb}, 20)

    # Arm64 SKUs (B2pts_v2, D4pls_v5 …) don't support Windows — exclude them
    if os_type == "Windows":
        docs = [d for d in docs
                if not re.search(r'Standard_[A-Za-z]\d+-?\d*p[a-z]', d.get("sku_name", ""), re.IGNORECASE)]

    # Exclude constrained-vCPU SKUs (e.g. E8-2ads_v7) whose active count < requested.
    # The AI Search index stores physical vCPUs; active_vcpu_count() gives the real number.
    from app.services.sql_pricing import active_vcpu_count as _active_vcpu_count
    docs = [d for d in docs
            if (_active_vcpu_count(d.get("sku_name", ""), d.get("vcpus") or 0) or d.get("vcpus") or 0) >= vcpus]

    if not docs:
        logger.warning(
            "bulk_pricing: no D/E/F/B SKU found for vcpus=%d ram_gb=%d — check AI Search index",
            vcpus, ram_gb,
        )
        _sku_match_cache[key] = []
        return []

    # Sort by tightest fit: fewest vCPUs, then smallest RAM, then no local NVMe
    # (prefer cheaper 'as'/'als' over 'ads'/'alds' when vcpu/RAM are equal),
    # then newest gen as final tiebreaker.
    def _has_local_nvme(sku_name: str) -> int:
        # 'd' in the suffix after the numeric vCPU count = local NVMe disk = pricier
        m = re.match(r'Standard_[A-Za-z]+\d+-?\d*([a-z]+)_v\d+', sku_name, re.IGNORECASE)
        return 1 if (m and 'd' in m.group(1).lower()) else 0

    def _sort_key(d: dict) -> tuple:
        m = re.search(r'_v(\d+)', d.get("sku_name", ""))
        gen = int(m.group(1)) if m else 1
        return (d.get("vcpus") or 999, d.get("ram_gb") or 999, _has_local_nvme(d.get("sku_name", "")), -gen)

    docs.sort(key=_sort_key)
    best = docs[0]

    matched_vcpus = best.get("vcpus") or vcpus
    matched_ram   = best.get("ram_gb") or ram_gb
    vcpu_excess   = matched_vcpus - vcpus
    if vcpu_excess > 16:
        logger.warning(
            "bulk_pricing: sanity — %s has vcpu_excess=%d for req vcpus=%d ram=%d",
            best.get("sku_name"), vcpu_excess, vcpus, ram_gb,
        )

    logger.info(
        "bulk_pricing: SKU match vcpus_req=%d ram_req=%d → %s (%d vCPU %d GB)",
        vcpus, ram_gb, best.get("sku_name"), matched_vcpus, matched_ram,
    )

    result = [best]
    _sku_match_cache[key] = result
    return result


# ── Cached VM price fetch ─────────────────────────────────────────────────────

async def _get_vm_prices(region: str, sku_name: str) -> list[dict]:
    """Fetch both Windows and Linux price items for a SKU, with caching.

    Filters out 'Basv2 Series Cloud Services' and similar items that share the
    same armSkuName/serviceName but have a different productName.  Those items
    are misclassified as Linux by detect_item_os and corrupt the Linux/Windows
    price split used to isolate the OS-license surcharge.
    """
    from app.services.azure_pricing import fetch_prices
    key = (region, sku_name)
    if key not in _vm_price_cache:
        raw = await fetch_prices(region, sku_name)
        _vm_price_cache[key] = [
            item for item in raw
            if "virtual machines" in (item.get("productName") or "").lower()
        ]
    return _vm_price_cache[key]


# ── Cached disk pricing ───────────────────────────────────────────────────────

async def _get_ssd_tier_price(region: str, tier: str) -> float:
    """Return the monthly Standard SSD price for a given tier in the region."""
    from app.services.azure_pricing import fetch_disk_tier_prices
    if region not in _ssd_price_cache:
        try:
            _ssd_price_cache[region] = await fetch_disk_tier_prices(region, "standard_ssd")
        except Exception as exc:
            logger.warning("bulk_pricing: SSD price fetch failed region=%s: %s", region, exc)
            _ssd_price_cache[region] = {}
    return _ssd_price_cache[region].get(tier, 0.0)


async def _get_premium_ssd_tier_price(region: str, tier: str) -> float:
    """Return the monthly Premium SSD (P-series) price for a given tier in the region."""
    from app.services.azure_pricing import fetch_disk_tier_prices
    if region not in _premium_ssd_cache:
        try:
            _premium_ssd_cache[region] = await fetch_disk_tier_prices(region, "premium_ssd")
        except Exception as exc:
            logger.warning("bulk_pricing: Premium SSD price fetch failed region=%s: %s", region, exc)
            _premium_ssd_cache[region] = {}
    return _premium_ssd_cache[region].get(tier, 0.0)


async def _get_pv2_monthly_cost(region: str, disk_gb: float) -> float:
    """Return the monthly Premium SSD v2 capacity cost for disk_gb GiB.

    Uses fetch_v2_capacity_rate() which returns $/GiB/month (already ×730).
    Covers capacity only — IOPS/throughput above the free baseline (3000/125)
    are not priced here, consistent with how the single-VM card prices Pv2.
    """
    from app.services.azure_pricing import fetch_v2_capacity_rate
    if region not in _pv2_rate_cache:
        try:
            _pv2_rate_cache[region] = await fetch_v2_capacity_rate(region)
        except Exception as exc:
            logger.warning("bulk_pricing: Pv2 rate fetch failed region=%s: %s", region, exc)
            _pv2_rate_cache[region] = None
    rate = _pv2_rate_cache[region]
    if rate is None:
        return 0.0
    return rate * max(1.0, disk_gb)


# ── Per-row price computation ─────────────────────────────────────────────────

async def _price_row(row: dict) -> dict:
    """Resolve SKU, fetch prices, and compute the full monthly breakdown for one VM row."""
    from app.utils.pricing_calculator import find_price, ri_monthly
    from app.services.sql_pricing import active_vcpu_count, sql_license_hourly

    region      = row["arm_region"]
    os_type     = row["os_type"]
    vcpus_req   = row["vcpus_req"]
    ram_gb_req  = row["ram_gb_req"]
    ssd_tier    = row["ssd_tier"]
    ssd_gib     = row["ssd_gib"]
    sql_billing = row["sql_billing"]
    win_ahb     = row.get("win_ahb", False)
    sql_ahb     = row.get("sql_ahb", False)
    disk_type   = row.get("disk_type", "standard_ssd")
    disk_gb_raw = row.get("disk_gb_raw", 0.0)

    # 1. Best-fit VM SKU (reuses advisor matching logic + caching)
    candidates = await _pick_sku_cached(region, os_type, vcpus_req, ram_gb_req)
    if not candidates:
        raise ValueError(
            f"no SKU found for {vcpus_req} vCPU / {ram_gb_req} GB RAM in {region} ({os_type})"
        )

    best = candidates[0]
    sku_name      = best["sku_name"]
    matched_vcpus = best.get("vcpus") or vcpus_req
    matched_ram   = best.get("ram_gb")

    # Active vCPU count — honours constrained-vCPU SKUs for SQL billing
    billing_vcpus = active_vcpu_count(sku_name, matched_vcpus) or matched_vcpus

    # 2. VM PAYG prices (Windows + Linux from same call)
    vm_items     = await _get_vm_prices(region, sku_name)
    win_item     = find_price(vm_items, "Windows", "Consumption")
    lin_item     = find_price(vm_items, "Linux",   "Consumption")
    windows_payg = win_item["retailPrice"] if win_item else None
    linux_payg   = lin_item["retailPrice"] if lin_item else None

    # 3. Disk monthly cost — routed through existing fetch functions by disk_type
    disk_gb_eff = disk_gb_raw if disk_gb_raw > 0 else 128.0
    if disk_type == "premium_ssd":
        p_tier, p_gib = resolve_premium_ssd_tier(disk_gb_eff)
        disk_monthly = await _get_premium_ssd_tier_price(region, p_tier)
        disk_label   = f"P-SSD {p_tier}/{p_gib} GiB"
    elif disk_type == "premium_ssd_v2":
        disk_monthly = await _get_pv2_monthly_cost(region, disk_gb_eff)
        disk_label   = f"Pv2 {int(disk_gb_eff)} GiB"
    else:  # standard_ssd (default)
        disk_monthly = await _get_ssd_tier_price(region, ssd_tier)
        disk_label   = f"E-SSD {ssd_tier}/{ssd_gib} GiB"

    # 4. Monthly breakdown
    #    compute  = base Linux-equivalent rate (bare metal, no OS charge)
    #    os_lic   = Windows Server OS surcharge (0 for Linux or Windows AHB)
    #    sql_lic  = SQL Server license per-vCPU (0 if no SQL or SQL AHB)
    if os_type == "Linux":
        compute_h = linux_payg or 0.0
        os_lic_h  = 0.0
    elif win_ahb:
        # Windows AHB: pay Linux compute rate, no OS license (BYOL)
        compute_h = linux_payg or 0.0
        os_lic_h  = 0.0
    else:
        # Windows License-Included: split into compute (Linux base) + OS surcharge
        if linux_payg is not None and windows_payg is not None:
            compute_h = linux_payg
            os_lic_h  = max(0.0, windows_payg - linux_payg)
        elif windows_payg is not None:
            # Only Windows price available — can't split; show all as compute
            compute_h = windows_payg
            os_lic_h  = 0.0
        else:
            compute_h = 0.0
            os_lic_h  = 0.0

    sql_h = (
        sql_license_hourly(billing_vcpus, sql_billing, sql_ahb=sql_ahb)
        if sql_billing else 0.0
    )

    monthly_compute    = round(compute_h * HOURS_PER_MONTH, 2)
    monthly_os_license = round(os_lic_h  * HOURS_PER_MONTH, 2)
    monthly_sql        = round(sql_h      * HOURS_PER_MONTH, 2)
    monthly_disk       = round(disk_monthly, 2)
    monthly_total      = round(monthly_compute + monthly_os_license + monthly_sql + monthly_disk, 2)

    # ── RI compute rates ──────────────────────────────────────────────────────
    # Linux RI items = compute-only rate (no OS license bundled).
    # OS license, SQL license, and disk remain at their PAYG rate for all tiers.
    ri1_item = find_price(vm_items, 'Linux', 'Reservation', '1 Year')
    ri3_item = find_price(vm_items, 'Linux', 'Reservation', '3 Years')

    if not ri1_item:
        logger.warning("bulk_pricing: no 1yr RI item for %s in %s — using PAYG compute", sku_name, region)
    if not ri3_item:
        logger.warning("bulk_pricing: no 3yr RI item for %s in %s — using PAYG compute", sku_name, region)

    monthly_compute_ri1 = round(ri_monthly(ri1_item), 2) if ri1_item else monthly_compute
    monthly_compute_ri3 = round(ri_monthly(ri3_item), 2) if ri3_item else monthly_compute
    monthly_total_ri1   = round(monthly_compute_ri1 + monthly_os_license + monthly_sql + monthly_disk, 2)
    monthly_total_ri3   = round(monthly_compute_ri3 + monthly_os_license + monthly_sql + monthly_disk, 2)

    return {
        # ── Input fields (pass-through) ───────────────────────────────────
        "sheet_row":    row["sheet_row"],
        "vm_name":      row["vm_name"],
        "region_label": row["region_label"],
        "role":         row["role"],
        "os_type":      os_type,
        "vcpus_req":    vcpus_req,
        "ram_gb_req":   ram_gb_req,
        "disk_gb_raw":  row["disk_gb_raw"],
        "sql_display":  row["sql_display"],
        "win_ahb":      win_ahb,
        "sql_ahb":      sql_ahb,
        "disk_type":    disk_type,
        # ── Resolved values ───────────────────────────────────────────────
        "sku_name":        sku_name,
        "matched_vcpus":   matched_vcpus,
        "matched_ram_gb":  matched_ram,
        "billing_vcpus":   billing_vcpus,
        "ssd_tier":        ssd_tier,
        "ssd_gib":         ssd_gib,
        "disk_label":      disk_label,
        "windows_payg_hr": windows_payg,
        "linux_payg_hr":   linux_payg,
        # ── Monthly costs ─────────────────────────────────────────────────
        "monthly_compute":     monthly_compute,
        "monthly_os_license":  monthly_os_license,
        "monthly_sql":         monthly_sql,
        "monthly_disk":        monthly_disk,
        "monthly_total":       monthly_total,
        "monthly_compute_ri1": monthly_compute_ri1,
        "monthly_compute_ri3": monthly_compute_ri3,
        "monthly_total_ri1":   monthly_total_ri1,
        "monthly_total_ri3":   monthly_total_ri3,
    }


# ── Top-level orchestrator ────────────────────────────────────────────────────

async def process_inventory(file_bytes: bytes) -> dict:
    """Parse an inventory Excel file and return full pricing breakdown.

    Returns a dict with:
        rows      — list of per-VM pricing dicts
        totals    — dict with monthly and annual grand totals
        warnings  — list of warning strings (parse errors, skipped rows, etc.)
        vm_count  — number of successfully priced VMs
    """
    rows, parse_warnings = parse_inventory_xlsx(file_bytes)

    result_rows: list[dict] = []
    pricing_warnings: list[str] = []

    for row in rows:
        try:
            priced = await _price_row(row)
            result_rows.append(priced)
        except Exception as exc:
            msg = f"'{row['vm_name']}' (row {row['sheet_row']}): pricing failed — {exc}"
            pricing_warnings.append(msg)
            logger.exception("bulk_pricing: %s", msg)

    monthly_compute    = sum(r["monthly_compute"]    for r in result_rows)
    monthly_os_license = sum(r["monthly_os_license"] for r in result_rows)
    monthly_sql        = sum(r["monthly_sql"]        for r in result_rows)
    monthly_disk       = sum(r["monthly_disk"]       for r in result_rows)
    monthly_total      = monthly_compute + monthly_os_license + monthly_sql + monthly_disk
    monthly_total_ri1  = sum(r["monthly_total_ri1"]  for r in result_rows)
    monthly_total_ri3  = sum(r["monthly_total_ri3"]  for r in result_rows)

    return {
        "rows":     result_rows,
        "totals": {
            "monthly_compute":    round(monthly_compute,    2),
            "monthly_os_license": round(monthly_os_license, 2),
            "monthly_sql":        round(monthly_sql,        2),
            "monthly_disk":       round(monthly_disk,       2),
            "monthly_total":      round(monthly_total,      2),
            "monthly_total_ri1":  round(monthly_total_ri1,  2),
            "monthly_total_ri3":  round(monthly_total_ri3,  2),
            "annual_total":       round(monthly_total * 12, 2),
            "annual_total_ri1":   round(monthly_total_ri1 * 12, 2),
            "annual_total_ri3":   round(monthly_total_ri3 * 12, 2),
        },
        "warnings": parse_warnings + pricing_warnings,
        "vm_count": len(result_rows),
    }


# ── Excel report generation ───────────────────────────────────────────────────

def generate_bulk_excel(result: dict) -> bytes:
    """Write the pricing breakdown to an xlsx file and return raw bytes."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from io import BytesIO

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "VM Pricing"

    # ── Styles ────────────────────────────────────────────────────────────────
    AZURE_BLUE   = "FF0078D4"
    HEADER_FONT  = Font(bold=True, color="FFFFFFFF", size=10)
    HEADER_FILL  = PatternFill(fill_type="solid", fgColor=AZURE_BLUE)
    TOTAL_FILL   = PatternFill(fill_type="solid", fgColor="FF004E8C")
    TOTAL_FONT   = Font(bold=True, color="FFFFFFFF", size=10)
    ALT_FILL     = PatternFill(fill_type="solid", fgColor="FFF0F5FF")
    WARN_FILL    = PatternFill(fill_type="solid", fgColor="FFFFF3CD")
    BORDER_SIDE  = Side(style="thin", color="FFD0D0D0")
    CELL_BORDER  = Border(bottom=Border(bottom=BORDER_SIDE).bottom)
    MONEY        = '#,##0.00'
    CENTER       = Alignment(horizontal="center", vertical="center")
    RIGHT        = Alignment(horizontal="right",  vertical="center")
    LEFT         = Alignment(horizontal="left",   vertical="center")

    def _hdr(cell, text):
        cell.value     = text
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = CENTER

    # ── Column layout ─────────────────────────────────────────────────────────
    cols = [
        ("#",                5),
        ("VM Name",         28),
        ("Region",          16),
        ("Resolved SKU",    24),
        ("vCPU",             6),
        ("RAM (GB)",         9),
        ("OS",              10),
        ("SQL Edition",     13),
        ("OS Disk",         12),
        ("Compute/mo",      12),
        ("OS Lic/mo",       11),
        ("SQL Lic/mo",      11),
        ("Disk/mo",         10),
        ("Total PAYG/mo",   13),
        ("1yr RI/mo",       12),
        ("3yr RI/mo",       12),
    ]
    for col_idx, (hdr, width) in enumerate(cols, start=1):
        cell = ws.cell(row=1, column=col_idx)
        _hdr(cell, hdr)
        ws.column_dimensions[cell.column_letter].width = width

    ws.row_dimensions[1].height = 18

    # ── Data rows ─────────────────────────────────────────────────────────────
    for i, row in enumerate(result["rows"], start=1):
        r = i + 1
        fill = ALT_FILL if i % 2 == 0 else None

        def _w(col, val, fmt=None, align=LEFT):
            c = ws.cell(row=r, column=col, value=val)
            if fill:
                c.fill = fill
            if fmt:
                c.number_format = fmt
            c.alignment = align
            return c

        disk_label = row.get("disk_label", f"E-SSD {row['ssd_tier']}/{row['ssd_gib']} GiB")
        os_label   = f"{row['os_type']} (AHB)" if row.get("win_ahb") else row["os_type"]
        sql_raw    = row["sql_display"]
        sql_label  = (f"{sql_raw} (AHB)" if row.get("sql_ahb") else sql_raw) if sql_raw else "—"
        _w(1,  i,                          align=CENTER)
        _w(2,  row["vm_name"])
        _w(3,  row["region_label"])
        _w(4,  row["sku_name"])
        _w(5,  row["matched_vcpus"],       align=CENTER)
        _w(6,  row.get("matched_ram_gb"),  align=CENTER)
        _w(7,  os_label,                   align=CENTER)
        _w(8,  sql_label,                  align=CENTER)
        _w(9,  disk_label,                 align=CENTER)
        _w(10, row["monthly_compute"],      MONEY, RIGHT)
        _w(11, row["monthly_os_license"],  MONEY, RIGHT)
        _w(12, row["monthly_sql"],         MONEY, RIGHT)
        _w(13, row["monthly_disk"],        MONEY, RIGHT)
        _w(14, row["monthly_total"],       MONEY, RIGHT)
        _w(15, row["monthly_total_ri1"],   MONEY, RIGHT)
        _w(16, row["monthly_total_ri3"],   MONEY, RIGHT)

    # ── Grand total row ───────────────────────────────────────────────────────
    total_row = len(result["rows"]) + 2
    t = result["totals"]

    def _tot(col, val, fmt=None):
        c = ws.cell(row=total_row, column=col, value=val)
        c.font      = TOTAL_FONT
        c.fill      = TOTAL_FILL
        c.alignment = RIGHT if fmt else CENTER
        if fmt:
            c.number_format = fmt
        return c

    _tot(1, "TOTAL",              fmt=None)
    for col in range(2, 10):
        c = ws.cell(row=total_row, column=col)
        c.fill = TOTAL_FILL
    _tot(10, t["monthly_compute"],    MONEY)
    _tot(11, t["monthly_os_license"], MONEY)
    _tot(12, t["monthly_sql"],        MONEY)
    _tot(13, t["monthly_disk"],       MONEY)
    _tot(14, t["monthly_total"],      MONEY)
    _tot(15, t["monthly_total_ri1"],  MONEY)
    _tot(16, t["monthly_total_ri3"],  MONEY)
    ws.row_dimensions[total_row].height = 18

    # ── Annual summary block ──────────────────────────────────────────────────
    ann_row = total_row + 2
    bold10  = Font(bold=True, size=10)
    _ann_labels = [
        ("Annual PAYG (×12)",    "annual_total"),
        ("Annual 1yr RI (×12)", "annual_total_ri1"),
        ("Annual 3yr RI (×12)", "annual_total_ri3"),
    ]
    for offset, (label, key) in enumerate(_ann_labels):
        r = ann_row + offset
        ws.cell(row=r, column=1, value=label).font = bold10
        c2 = ws.cell(row=r, column=14, value=t[key])
        c2.number_format = MONEY
        c2.font          = bold10
        c2.alignment     = RIGHT

    note_row = ann_row + len(_ann_labels) + 1
    ws.cell(row=note_row, column=1,
            value=(
                "Note: OS disk only (no data disks priced). PAYG rates. "
                "Windows/SQL AHB applied per row where specified. "
                "Standard SSD LRS by default; Premium SSD and Premium SSD v2 available per row. "
                "RI discount on compute only — OS/SQL license at PAYG rate."
            )
    ).font = Font(italic=True, size=9, color="FF666666")

    # ── Warnings sheet ────────────────────────────────────────────────────────
    if result["warnings"]:
        ws2 = wb.create_sheet("Warnings")
        ws2.column_dimensions["A"].width = 100
        ws2.cell(row=1, column=1, value="Warnings / Skipped rows").font = Font(bold=True, size=11)
        for i, msg in enumerate(result["warnings"], start=2):
            ws2.cell(row=i, column=1, value=msg)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
