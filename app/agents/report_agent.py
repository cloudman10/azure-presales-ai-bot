"""
app/agents/report_agent.py

Generate Excel (.xlsx) and PDF reports from Azure VM pricing text (single-card)
and from the structured numeric basket model (multi-item quote).
"""

import io
from datetime import datetime

_TYPE_NAMES = {
    'premium_ssd':    'Premium SSD',
    'standard_ssd':   'Standard SSD',
    'standard_hdd':   'Standard HDD',
    'premium_ssd_v2': 'Premium SSD v2',
}

def _size_str(size_gb: int) -> str:
    return f"{size_gb // 1024} TiB" if size_gb >= 1024 else f"{size_gb} GiB"


# ── Excel ──────────────────────────────────────────────────────────────────────

def generate_excel(pricing_text: str, session_id: str = "") -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "VM Pricing Summary"

    # ── Branding header ────────────────────────────────────────────────────────
    ws.merge_cells("A1:D1")
    ws["A1"] = "HyperXen.ai — Azure VM Pricing Report"
    ws["A1"].font      = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill      = PatternFill("solid", fgColor="0078D4")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A2:D2")
    ws["A2"] = f"Generated: {datetime.now().strftime('%d %b %Y %H:%M')}"
    ws["A2"].font      = Font(italic=True, size=10, color="8B95A2")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18

    # ── Parse pricing text into rows ───────────────────────────────────────────
    row = 4
    for line in pricing_text.splitlines():
        stripped = line.strip()

        if not stripped:
            row += 1
            continue

        if stripped.startswith("===") and stripped.endswith("==="):
            # Main section title (e.g. === Azure VM Pricing Estimate ===)
            label = stripped.strip("=").strip()
            ws.merge_cells(f"A{row}:D{row}")
            cell = ws.cell(row=row, column=1, value=label)
            cell.font      = Font(bold=True, size=11, color="FFFFFF")
            cell.fill      = PatternFill("solid", fgColor="005A9E")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[row].height = 22

        elif stripped.startswith("---") and stripped.endswith("---"):
            # Sub-section header (e.g. --- PAYG ---)
            label = stripped.strip("-").strip()
            ws.merge_cells(f"A{row}:D{row}")
            cell = ws.cell(row=row, column=1, value=label)
            cell.font      = Font(bold=True, size=10, color="005A9E")
            cell.fill      = PatternFill("solid", fgColor="E8F4FD")
            cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            ws.row_dimensions[row].height = 18

        elif ":" in stripped and not stripped.startswith(" "):
            # Key: value pair → split across two columns
            key, _, val = stripped.partition(":")
            kc = ws.cell(row=row, column=1, value=key.strip())
            kc.font = Font(bold=True, size=10)
            vc = ws.cell(row=row, column=2, value=val.strip())
            vc.font = Font(size=10)

        else:
            # Indented detail line — merge across all columns
            ws.merge_cells(f"A{row}:D{row}")
            cell = ws.cell(row=row, column=1, value=stripped)
            cell.font      = Font(size=10, color="444444")
            cell.alignment = Alignment(indent=2)

        row += 1

    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── PDF ────────────────────────────────────────────────────────────────────────

def generate_pdf(pricing_text: str, session_id: str = "") -> bytes:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, HRFlowable
    from reportlab.lib.enums import TA_CENTER, TA_LEFT

    AZURE_BLUE = colors.HexColor("#0078D4")
    DARK_BLUE  = colors.HexColor("#005A9E")
    LIGHT_BLUE = colors.HexColor("#E8F4FD")
    GREY       = colors.HexColor("#8B95A2")
    DARK       = colors.HexColor("#1A1A2E")

    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        "RPTitle", parent=styles["Normal"],
        fontSize=16, fontName="Helvetica-Bold",
        textColor=colors.white, backColor=AZURE_BLUE,
        alignment=TA_CENTER, spaceAfter=0,
        borderPadding=(10, 14, 10, 14),
    )
    sub_style = ParagraphStyle(
        "RPSub", parent=styles["Normal"],
        fontSize=9, fontName="Helvetica-Oblique",
        textColor=GREY, alignment=TA_CENTER, spaceAfter=14,
    )
    main_title_style = ParagraphStyle(
        "RPMainTitle", parent=styles["Normal"],
        fontSize=11, fontName="Helvetica-Bold",
        textColor=colors.white, backColor=DARK_BLUE,
        alignment=TA_CENTER, spaceBefore=8, spaceAfter=4,
        borderPadding=(5, 10, 5, 10),
    )
    section_style = ParagraphStyle(
        "RPSection", parent=styles["Normal"],
        fontSize=10, fontName="Helvetica-Bold",
        textColor=DARK_BLUE, backColor=LIGHT_BLUE,
        spaceBefore=10, spaceAfter=3,
        borderPadding=(4, 8, 4, 8),
    )
    kv_style = ParagraphStyle(
        "RPKV", parent=styles["Normal"],
        fontSize=9, fontName="Helvetica",
        textColor=DARK, spaceBefore=2, spaceAfter=2, leftIndent=8,
    )
    body_style = ParagraphStyle(
        "RPBody", parent=styles["Normal"],
        fontSize=9, fontName="Helvetica",
        textColor=DARK, spaceBefore=1, spaceAfter=1, leftIndent=14,
    )
    footer_style = ParagraphStyle(
        "RPFooter", parent=styles["Normal"],
        fontSize=8, textColor=GREY,
        alignment=TA_CENTER, spaceBefore=6,
    )

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=20*mm, rightMargin=20*mm,
        topMargin=20*mm, bottomMargin=20*mm,
    )

    story = [
        Paragraph("HyperXen.ai — Azure VM Pricing Report", title_style),
        Paragraph(f"Generated: {datetime.now().strftime('%d %b %Y %H:%M')}", sub_style),
    ]

    for line in pricing_text.splitlines():
        stripped = line.strip()

        if not stripped:
            story.append(Spacer(1, 4))
            continue

        if stripped.startswith("===") and stripped.endswith("==="):
            label = stripped.strip("=").strip()
            story.append(Paragraph(label, main_title_style))

        elif stripped.startswith("---") and stripped.endswith("---"):
            label = stripped.strip("-").strip()
            story.append(Paragraph(label, section_style))

        elif ":" in stripped and not stripped.startswith(" "):
            key, _, val = stripped.partition(":")
            story.append(Paragraph(
                f"<b>{key.strip()}:</b> {val.strip()}", kv_style
            ))

        else:
            story.append(Paragraph(stripped, body_style))

    story += [
        Spacer(1, 10),
        HRFlowable(width="100%", thickness=0.5, color=GREY),
        Paragraph(
            "All prices are Microsoft Retail RRP. CSP pricing would be lower.",
            footer_style,
        ),
    ]

    doc.build(story)
    return buf.getvalue()


# ── Basket Excel (structured numeric) ──────────────────────────────────────────

def generate_excel_basket(items: list[dict], grand_total: float) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "Azure VM Quote"

    ws.column_dimensions["A"].width = 44
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 8
    ws.column_dimensions["D"].width = 16

    # Branding header
    ws.merge_cells("A1:D1")
    ws["A1"] = "HyperXen.ai — Azure VM Quote"
    ws["A1"].font      = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill      = PatternFill("solid", fgColor="0078D4")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A2:D2")
    count_label = f"{len(items)} item{'s' if len(items) != 1 else ''}"
    ws["A2"] = f"Generated: {datetime.now().strftime('%d %b %Y %H:%M')}  |  {count_label}"
    ws["A2"].font      = Font(italic=True, size=10, color="D4E8FF")
    ws["A2"].fill      = PatternFill("solid", fgColor="0078D4")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 20

    # Column headers
    row = 4
    hdr_fill = PatternFill("solid", fgColor="005A9E")
    for col, label in enumerate(["Description", "Unit $/mo", "x", "Extended $/mo"], 1):
        c = ws.cell(row=row, column=col, value=label)
        c.font      = Font(bold=True, size=9, color="FFFFFF")
        c.fill      = hdr_fill
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 18
    row += 1

    for item in items:
        # Item section header
        ws.merge_cells(f"A{row}:D{row}")
        hdr_text = f"{item['count']}x {item['sku']}  |  {item['os']}  |  {item['region']}"
        c = ws.cell(row=row, column=1, value=hdr_text)
        c.font      = Font(bold=True, size=10, color="FFFFFF")
        c.fill      = PatternFill("solid", fgColor="1A5494")
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[row].height = 20
        row += 1

        # VM row
        ws.cell(row=row, column=1, value=f"VM ({item['term']}/month)").font = Font(size=10)
        c = ws.cell(row=row, column=2, value=item["vm_unit_cost"])
        c.number_format = '"$"#,##0.00'
        ws.cell(row=row, column=3, value=f"x{item['count']}").font = Font(size=10, color="5A7A9A")
        c = ws.cell(row=row, column=4, value=round(item["vm_unit_cost"] * item["count"], 2))
        c.number_format = '"$"#,##0.00'
        c.font = Font(size=10, bold=True)
        row += 1

        # Disk rows
        for disk in item.get("disks", []):
            d_type = _TYPE_NAMES.get(disk.get("type", ""), disk.get("type", ""))
            d_size = _size_str(disk.get("size_gb", 0))
            d_tier = disk.get("tier", "")
            d_role = disk.get("role", "Disk")
            ws.cell(row=row, column=1, value=f"  {d_role} — {d_type} {d_tier} ({d_size})").font = Font(size=9, color="444444")
            c = ws.cell(row=row, column=2, value=disk.get("cost", 0))
            c.number_format = '"$"#,##0.00'
            c.font = Font(size=9, color="444444")
            ws.cell(row=row, column=3, value=f"x{item['count']}").font = Font(size=9, color="5A7A9A")
            c = ws.cell(row=row, column=4, value=round(disk.get("cost", 0) * item["count"], 2))
            c.number_format = '"$"#,##0.00'
            c.font = Font(size=9, color="444444")
            row += 1

        # Line total
        ws.merge_cells(f"A{row}:C{row}")
        ws.cell(row=row, column=1, value="Line Total").font = Font(bold=True, size=10)
        c = ws.cell(row=row, column=4, value=item["line_total"])
        c.number_format = '"$"#,##0.00'
        c.font = Font(bold=True, size=11, color="0078D4")
        row += 2  # blank separator

    # Grand total
    ws.merge_cells(f"A{row}:C{row}")
    c = ws.cell(row=row, column=1, value="GRAND TOTAL")
    c.font      = Font(bold=True, size=12, color="FFFFFF")
    c.fill      = PatternFill("solid", fgColor="0078D4")
    c.alignment = Alignment(horizontal="right", vertical="center", indent=1)
    c = ws.cell(row=row, column=4, value=grand_total)
    c.number_format = '"$"#,##0.00'
    c.font      = Font(bold=True, size=12, color="FFFFFF")
    c.fill      = PatternFill("solid", fgColor="0078D4")
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 24
    row += 2

    # Footnotes
    for note in [
        "* Standard SSD/HDD: capacity cost only; transaction costs excluded.",
        "All prices are Microsoft Retail RRP. Monthly estimate = 730 hrs.",
        "Prices fetched from Azure Retail Prices API at time of query.",
    ]:
        ws.merge_cells(f"A{row}:D{row}")
        ws.cell(row=row, column=1, value=note).font = Font(size=8, color="8B95A2", italic=True)
        row += 1

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Basket PDF (structured numeric) ────────────────────────────────────────────

def generate_pdf_basket(items: list[dict], grand_total: float) -> bytes:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, HRFlowable, Table, TableStyle,
    )
    from reportlab.lib.enums import TA_CENTER

    AZURE_BLUE = colors.HexColor("#0078D4")
    DARK_BLUE  = colors.HexColor("#1A5494")
    LIGHT_BLUE = colors.HexColor("#E8F4FD")
    GREY       = colors.HexColor("#8B95A2")
    DARK       = colors.HexColor("#1A1A2E")

    styles = getSampleStyleSheet()

    item_hdr_style = ParagraphStyle(
        "BQItemHdr", parent=styles["Normal"],
        fontSize=10, fontName="Helvetica-Bold",
        textColor=colors.white, backColor=DARK_BLUE,
        spaceBefore=12, spaceAfter=2,
        borderPadding=(5, 8, 5, 8),
    )
    footer_style = ParagraphStyle(
        "BQFooter", parent=styles["Normal"],
        fontSize=8, textColor=GREY,
        alignment=TA_CENTER, spaceBefore=4,
    )

    # A4 text width = 170mm; column split: desc | unit | mult | ext
    col_w = [85*mm, 30*mm, 15*mm, 30*mm]

    base_ts = [
        ("BACKGROUND", (0, 0), (-1, 0), LIGHT_BLUE),
        ("TEXTCOLOR",  (0, 0), (-1, 0), DARK_BLUE),
        ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",   (0, 0), (-1, 0), 8),
        ("FONTSIZE",   (0, 1), (-1, -1), 9),
        ("ALIGN",      (1, 0), (-1, -1), "RIGHT"),
        ("ALIGN",      (2, 0), (2, -1), "CENTER"),
        ("TOPPADDING",    (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FBFF")]),
        ("LINEBELOW",  (0, -1), (-1, -1), 0.5, DARK_BLUE),
    ]

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=20*mm, rightMargin=20*mm,
        topMargin=20*mm, bottomMargin=20*mm,
    )

    count_label = f"{len(items)} item{'s' if len(items) != 1 else ''}"
    date_str    = datetime.now().strftime('%d %b %Y %H:%M')
    # Two-row header table — both rows share the blue band so title and subtitle
    # are guaranteed to be stacked with no overlap and with readable contrast.
    hdr_w = A4[0] - 40*mm   # full text-area width (page width - L+R margins)
    hdr_tbl = Table(
        [
            ["HyperXen.ai — Azure VM Quote"],
            [f"Generated: {date_str}  │  {count_label}"],
        ],
        colWidths=[hdr_w],
    )
    hdr_tbl.setStyle(TableStyle([
        ("BACKGROUND",    (0, 0), (-1, -1), AZURE_BLUE),
        ("TEXTCOLOR",     (0, 0), (0, 0), colors.white),
        ("FONTNAME",      (0, 0), (0, 0), "Helvetica-Bold"),
        ("FONTSIZE",      (0, 0), (0, 0), 16),
        ("TEXTCOLOR",     (0, 1), (0, 1), colors.HexColor("#D4E8FF")),
        ("FONTNAME",      (0, 1), (0, 1), "Helvetica-Oblique"),
        ("FONTSIZE",      (0, 1), (0, 1), 10),
        ("ALIGN",         (0, 0), (-1, -1), "CENTER"),
        ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING",    (0, 0), (0, 0), 14),
        ("BOTTOMPADDING", (0, 0), (0, 0), 6),
        ("TOPPADDING",    (0, 1), (0, 1), 6),
        ("BOTTOMPADDING", (0, 1), (0, 1), 14),
    ]))
    story = [hdr_tbl, Spacer(1, 14)]

    for item in items:
        hdr_text = f"{item['count']}× {item['sku']}  |  {item['os']}  |  {item['region']}"
        story.append(Paragraph(hdr_text, item_hdr_style))

        tdata = [["Description", "Unit $/mo", "×", "Extended $/mo"]]

        vm_ext = round(item["vm_unit_cost"] * item["count"], 2)
        tdata.append([
            f"VM ({item['term']}/month)",
            f"${item['vm_unit_cost']:.2f}",
            f"×{item['count']}",
            f"${vm_ext:.2f}",
        ])

        for disk in item.get("disks", []):
            d_type = _TYPE_NAMES.get(disk.get("type", ""), disk.get("type", ""))
            d_size = _size_str(disk.get("size_gb", 0))
            d_tier = disk.get("tier", "")
            d_role = disk.get("role", "Disk")
            d_ext  = round(disk.get("cost", 0) * item["count"], 2)
            tdata.append([
                f"  {d_role} — {d_type} {d_tier} ({d_size})",
                f"${disk.get('cost', 0):.2f}",
                f"×{item['count']}",
                f"${d_ext:.2f}",
            ])

        # Line total row (spans first three cols)
        last = len(tdata)
        tdata.append(["Line Total", "", "", f"${item['line_total']:.2f}"])

        ts = TableStyle(base_ts + [
            ("SPAN",      (0, last), (2, last)),
            ("FONTNAME",  (0, last), (-1, last), "Helvetica-Bold"),
            ("FONTSIZE",  (0, last), (-1, last), 10),
            ("TEXTCOLOR", (3, last), (3, last), AZURE_BLUE),
            ("LINEABOVE", (0, last), (-1, last), 0.5, GREY),
        ])

        story.append(Table(tdata, colWidths=col_w, style=ts))

    # Grand total
    story.append(Spacer(1, 8))
    gt_data = [["GRAND TOTAL", "", "", f"${grand_total:.2f}"]]
    gt_ts = TableStyle([
        ("BACKGROUND",    (0, 0), (-1, -1), AZURE_BLUE),
        ("TEXTCOLOR",     (0, 0), (-1, -1), colors.white),
        ("FONTNAME",      (0, 0), (-1, -1), "Helvetica-Bold"),
        ("FONTSIZE",      (0, 0), (-1, -1), 12),
        ("SPAN",          (0, 0), (2, 0)),
        ("ALIGN",         (3, 0), (3, 0), "RIGHT"),
        ("TOPPADDING",    (0, 0), (-1, -1), 7),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
    ])
    story.append(Table(gt_data, colWidths=col_w, style=gt_ts))
    story.append(Spacer(1, 10))
    story.append(HRFlowable(width="100%", thickness=0.5, color=GREY))
    story += [
        Paragraph("* Standard SSD/HDD: capacity cost only; transaction costs excluded.", footer_style),
        Paragraph("All prices are Microsoft Retail RRP. Monthly estimate = 730 hrs.", footer_style),
        Paragraph("Prices fetched from Azure Retail Prices API at time of query.", footer_style),
    ]

    doc.build(story)
    return buf.getvalue()
